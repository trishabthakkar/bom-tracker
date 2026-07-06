import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_FIELDS = {"part_number"}
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}

COLUMN_ALIASES = {
    "part_number": {
        "part number",
        "part no",
        "part no.",
        "part_number",
        "partnumber",
        "part",
        "item number",
        "item no",
        "item",
        "sku",
    },
    "description": {
        "description",
        "desc",
        "part description",
        "item description",
        "name",
    },
    "parent_assembly": {
        "parent assembly",
        "parent",
        "parent part",
        "parent part number",
        "parent_part_number",
        "assembly",
        "assembly number",
    },
    "child_assembly": {
        "child assembly",
        "child",
        "child part",
        "child part number",
        "child_part_number",
        "component",
        "component part",
    },
    "revision": {
        "revision",
        "rev",
        "part revision",
        "revision level",
        "version",
    },
}


class BomParserError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedBomItem:
    row_number: int
    part_number: str
    description: str | None
    parent_assembly: str | None
    child_assembly: str | None
    revision: str | None


@dataclass(frozen=True)
class BomParseResult:
    rows: list[ParsedBomItem]


def parse_bom_file(path: str | Path) -> BomParseResult:
    file_path = Path(path)
    extension = file_path.suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise BomParserError("Only CSV and XLSX BOM files can be parsed.")

    raw_rows = _read_csv(file_path) if extension == ".csv" else _read_xlsx(file_path)
    return _parse_rows(raw_rows)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as input_file:
        reader = csv.DictReader(input_file)
        return [dict(row) for row in reader]


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = ["" if value is None else str(value) for value in rows[0]]
    parsed_rows: list[dict[str, Any]] = []

    for values in rows[1:]:
        parsed_rows.append(
            {
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
            }
        )

    return parsed_rows


def _parse_rows(raw_rows: list[dict[str, Any]]) -> BomParseResult:
    if not raw_rows:
        raise BomParserError("BOM file does not contain any data rows.")

    column_map = _build_column_map(list(raw_rows[0].keys()))
    missing_fields = sorted(REQUIRED_FIELDS - set(column_map.keys()))

    if missing_fields:
        missing = ", ".join(missing_fields)
        raise BomParserError(f"BOM file is missing required columns: {missing}.")

    parsed_rows: list[ParsedBomItem] = []

    for index, raw_row in enumerate(raw_rows, start=2):
        part_number = _clean_value(raw_row.get(column_map["part_number"]))
        if not part_number:
            continue

        parsed_rows.append(
            ParsedBomItem(
                row_number=index,
                part_number=part_number,
                description=_clean_optional(raw_row, column_map, "description"),
                parent_assembly=_clean_optional(raw_row, column_map, "parent_assembly"),
                child_assembly=_clean_optional(raw_row, column_map, "child_assembly"),
                revision=_clean_optional(raw_row, column_map, "revision"),
            )
        )

    if not parsed_rows:
        raise BomParserError("BOM file does not contain any rows with part numbers.")

    return BomParseResult(rows=parsed_rows)


def _build_column_map(headers: list[str]) -> dict[str, str]:
    normalized_headers = {_normalize_header(header): header for header in headers if header}
    column_map: dict[str, str] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if normalized_alias in normalized_headers:
                column_map[field] = normalized_headers[normalized_alias]
                break

    return column_map


def _normalize_header(value: str) -> str:
    return "".join(character for character in value.lower().strip() if character.isalnum())


def _clean_optional(
    raw_row: dict[str, Any],
    column_map: dict[str, str],
    field: str,
) -> str | None:
    column = column_map.get(field)
    return _clean_value(raw_row.get(column)) if column else None


def _clean_value(value: Any) -> str | None:
    if value is None:
        return None

    cleaned = str(value).strip()
    return cleaned or None
